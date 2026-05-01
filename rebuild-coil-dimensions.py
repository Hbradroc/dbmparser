# Regenerate coilsDimensionsData.json + coilsDimensionsData.js from drawings/*.xlsx
# Requires: pip install openpyxl
# Run: python rebuild-coil-dimensions.py
# Uses network (GitHub raw) unless --local and ./drawings has the files.

from __future__ import annotations

import argparse
import io
import json
import sys
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
DRAWINGS_ROOT = SCRIPT_DIR / "drawings"
INDEX_JSON = SCRIPT_DIR / "coils-drawings-index.json"
OUT_JSON = SCRIPT_DIR / "coilsDimensionsData.json"
OUT_JS = SCRIPT_DIR / "coilsDimensionsData.js"


def gh_raw_url(rel_path: str) -> str:
    from urllib.parse import quote

    parts = rel_path.replace("\\", "/").split("/")
    return (
        "https://raw.githubusercontent.com/Hbradroc/dbmparser/main/drawings/"
        + "/".join(quote(p, safe="/") for p in parts)
    )


def fetch_bytes(url: str) -> bytes:
    from urllib.request import Request, urlopen

    req = Request(url, headers={"User-Agent": "dbmparser-rebuild-dimensions/1"})
    with urlopen(req, timeout=120) as r:
        return r.read()


def load_workbook(rel_path: str, use_local: bool):
    rp = Path(rel_path.replace("\\", "/"))
    local_file = DRAWINGS_ROOT / rp
    if use_local and local_file.is_file():
        return openpyxl.load_workbook(local_file, read_only=True, data_only=True)
    url = gh_raw_url(str(rp))
    data = fetch_bytes(url)
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def norm_cell(c):
    if c is None:
        return ""
    if isinstance(c, datetime):
        return c.isoformat(sep=" ", timespec="minutes")
    if isinstance(c, (date, time)):
        return c.isoformat()
    return str(c).replace("\n", " ").strip()


def json_cell(v):
    if isinstance(v, datetime):
        return v.isoformat(sep=" ", timespec="minutes")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def _row_cell_tokens(row, n=16):
    return [norm_cell(x).upper() for x in (row or ())[:n]]


def _row_has_substr(cells, needle):
    return any(needle in c for c in cells if c)


def find_gx_header_row(ws, max_scan=55):
    for ri, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        if not row:
            continue
        up = _row_cell_tokens(row)
        joined = "|".join(up)
        has_unit = "UNIT" in up or "UNIT" in joined
        # Headers like "L1 aleta" / "T1 aleta" (substring match, not whole cell).
        has_l1 = _row_has_substr(up, "L1")
        has_t1 = _row_has_substr(up, "T1")
        has_t2 = _row_has_substr(up, "T2")
        if has_unit and has_l1 and has_t1:
            return ri
        if has_unit and has_l1 and has_t2:
            return ri
        if row[0] and norm_cell(row[0]).strip().upper() == "UNIT":
            return ri
    return None


def find_dvh_header_row(ws):
    for ri, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
        if not row:
            continue
        up = [norm_cell(x).upper() for x in row[:18]]
        joined = "|".join(up)
        if "DVH-W-Y" in joined or norm_cell(row[0]).upper() == "DVH-W-Y":
            return ri
        if ("L1" in up or "MEASURE TO DECREASE" in joined) and "T1" in up and "T2" in up:
            return ri
    return None


def trim_table(headers, rows, max_cols=28):
    if not headers:
        return headers, rows
    n = len(headers)
    while n > 2 and not str(headers[n - 1]).strip():
        n -= 1
        headers = headers[:n]
        rows = [r[:n] for r in rows]
    rows = rows[:280]
    return headers, rows


def extract_sheet(wb, sheet_name: str, max_cols=32):
    ws = wb[sheet_name]

    gx_hr = find_gx_header_row(ws)
    if gx_hr is not None:
        headers = []
        hdr_len = max_cols
        out_rows = []
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri < gx_hr:
                continue
            if ri == gx_hr:
                rl = list((tuple(row) + (None,) * max_cols)[:max_cols])
                hdr_len = max_cols
                headers = [
                    norm_cell(c) if norm_cell(c) else f"C{j}"
                    for j, c in enumerate(rl)
                ]
                continue
            rl = [
                json_cell(v)
                for v in list((tuple(row) + (None,) * hdr_len)[:hdr_len])
            ]
            if all(v is None or str(v).strip() == "" for v in rl[:5]):
                continue
            out_rows.append(rl)
            if len(out_rows) > 300:
                break
        headers, out_rows = trim_table(headers, out_rows)
        return {"layout": "gx", "headers": headers, "rows": out_rows}

    dvh_hr = find_dvh_header_row(ws)
    if dvh_hr is None:
        return None

    headers = []
    hdr_len = 0
    out_rows = []
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri < dvh_hr:
            continue
        if ri == dvh_hr:
            hdr_len = 24
            rl = list((tuple(row) + (None,) * hdr_len)[:hdr_len])
            headers = [
                norm_cell(c) if norm_cell(c) else f"C{j}"
                for j, c in enumerate(rl)
            ]
            continue
        rl = [
            json_cell(v)
            for v in list((tuple(row) + (None,) * hdr_len)[:hdr_len])
        ]
        if all(v is None or str(v).strip() == "" for v in rl[:8]):
            continue
        out_rows.append(rl)
        if len(out_rows) > 260:
            break
    headers, out_rows = trim_table(headers, out_rows)
    return {"layout": "dvh", "headers": headers, "rows": out_rows}


def norm_header_key(headers):
    return tuple(norm_cell(h).upper().strip() for h in headers or [])


def merge_gx_blocks(blocks: list):
    """Append rows from every GX sheet that shares the same header row (e.g. Foglio 2 with GX11)."""
    if not blocks:
        return None
    base = blocks[0]
    hdr_key = norm_header_key(base["headers"])
    merged_rows = list(base["rows"])
    sheet_parts = [str(base.get("sheetName") or "")]
    for b in blocks[1:]:
        if norm_header_key(b["headers"]) != hdr_key:
            print(
                f"  skip merge {base.get('sheetName')} + {b.get('sheetName')}: headers differ",
                file=sys.stderr,
            )
            continue
        merged_rows.extend(b["rows"])
        sheet_parts.append(str(b.get("sheetName") or ""))
    merged = dict(base)
    merged["rows"] = merged_rows
    merged["sheetName"] = ", ".join(s for s in sheet_parts if s)
    return merged


def process_workbook(rel_path: str, use_local: bool) -> dict | None:
    wb = load_workbook(rel_path, use_local)
    try:
        gx_blocks = []
        first_other = None
        for name in wb.sheetnames:
            block = extract_sheet(wb, name)
            if not block or not block.get("rows"):
                continue
            block["sheetName"] = name
            if block.get("layout") == "gx":
                gx_blocks.append(block)
            elif first_other is None:
                first_other = block
        if gx_blocks:
            merged = merge_gx_blocks(gx_blocks) if len(gx_blocks) > 1 else gx_blocks[0]
            merged["relPath"] = rel_path.replace("\\", "/")
            return merged
        if first_other:
            first_other["relPath"] = rel_path.replace("\\", "/")
            return first_other
    except Exception as e:
        print(f"  skip (error) {rel_path}: {e}", file=sys.stderr)
        return None
    finally:
        wb.close()
    return None


def apply_last_wins_per_app(nested_updates: dict[str, dict], geom: str, app: str, table: dict):
    """Keep one workbook per geom+app — later files overwrite (often REV01 wins if sorted)."""
    nested_updates.setdefault(geom, {})[app] = table


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--local",
        action="store_true",
        help="Prefer local drawings/ copies",
    )
    args = ap.parse_args()

    if not INDEX_JSON.is_file():
        print("Missing coils-drawings-index.json", file=sys.stderr)
        sys.exit(1)

    index_entries = json.loads(INDEX_JSON.read_text(encoding="utf-8-sig"))
    xlsx_sorted = sorted(
        [
            e
            for e in index_entries
            if str(e.get("ext", "")).lower() == ".xlsx"
            and str(e.get("geometry", "")) in ("P25", "P3012", "P40")
        ],
        key=lambda x: str(x.get("relPath", "")),
    )

    nested: dict = {}
    for e in xlsx_sorted:
        geom = str(e["geometry"])
        app = str(e["application"])
        rel_path = str(e["relPath"]).replace("\\", "/")
        print(rel_path, flush=True)
        table = process_workbook(rel_path, args.local)
        if table:
            apply_last_wins_per_app(nested, geom, app, table)

    OUT_JSON.write_text(json.dumps(nested, indent=2, ensure_ascii=False), encoding="utf-8")
    OUT_JS.write_text(
        "window.DBMM_COIL_DIM = "
        + json.dumps(nested, separators=(",", ":"), ensure_ascii=False)
        + ";\n",
        encoding="utf-8",
    )
    stats = ", ".join(f"{g}:{len(nested[g])}" for g in sorted(nested))
    print("OK", OUT_JSON.name, OUT_JS.name, stats)


if __name__ == "__main__":
    main()
